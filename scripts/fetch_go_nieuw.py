#!/usr/bin/env python3
"""
Download en parseer het GO! nieuw leerplan basisonderwijs (12 doelensets).

Stappen:
  1. Scrape de GO! Pro-pagina voor Excel- en PDF-downloadlinks.
  2. Download alle 12 doelenset-Excel-bestanden.
  3. Parseer doelzin-rijen naar JSONL.
  4. Sla vakspecifieke visieteksten (PDF) op.

Gebruik:
  python3 scripts/fetch_go_nieuw.py
  python3 scripts/fetch_go_nieuw.py --skip-download  # alleen herparsen

Installatie:
  pip install -r scripts/requirements-curriculum.txt
"""

from __future__ import annotations

import argparse
import json
import logging
import re
import sys
from collections import Counter
from datetime import datetime, timezone
from pathlib import Path
from typing import Any
from urllib.parse import urljoin

import openpyxl
import requests
from bs4 import BeautifulSoup

BASE_URL = "https://pro.g-o.be"
PAGE_URL = (
    f"{BASE_URL}/themas/leerplannen/basisonderwijs/nieuw-leerplan-basisonderwijs/"
)
OUTPUT_PATH = Path(__file__).resolve().parents[1] / "data/go_nieuw/go_nieuw_volledig.jsonl"
REPORT_PATH = OUTPUT_PATH.with_name("go_nieuw_scrape_report.json")
METADATA_PATH = OUTPUT_PATH.with_suffix(OUTPUT_PATH.suffix + ".meta.json")
DATA_DIR = OUTPUT_PATH.parent
EXCEL_DIR = DATA_DIR / "excel"
PDF_DIR = DATA_DIR / "pdf"

COMPONENT_MAP = {
    "begrijpen": "Begrijpen",
    "gebruiken": "Gebruiken",
    "engageren": "Engageren",
}

logger = logging.getLogger("fetch_go_nieuw")


def configure_logging(verbose: bool) -> None:
    logging.basicConfig(
        level=logging.DEBUG if verbose else logging.INFO,
        format="%(asctime)s %(levelname)s %(message)s",
        datefmt="%H:%M:%S",
    )


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value)).strip()


def normalize_component(value: Any) -> str:
    raw = clean_text(value).casefold()
    return COMPONENT_MAP.get(raw, clean_text(value))


def discipline_from_label(label: str) -> str:
    text = clean_text(label)
    text = re.sub(r"^Doelenset BaO\s+", "", text, flags=re.I)
    return text


class GoNieuwFetcher:
    def __init__(
        self,
        *,
        output: Path = OUTPUT_PATH,
        skip_download: bool = False,
    ) -> None:
        self.output = output
        self.report_path = output.with_name("go_nieuw_scrape_report.json")
        self.metadata_path = output.with_suffix(output.suffix + ".meta.json")
        self.skip_download = skip_download
        self.session = requests.Session()
        self.session.headers.update(
            {
                "User-Agent": (
                    "Mozilla/5.0 (compatible; Leerkrachtentools-GO-scraper/1.0)"
                ),
                "Accept": "*/*",
            }
        )
        self.report: dict[str, Any] = {
            "excel_downloads": [],
            "pdf_downloads": [],
            "warnings": [],
            "discipline_counts": {},
            "started_at": datetime.now(timezone.utc).isoformat(),
        }

    def run(self) -> None:
        page_html = self._fetch_page()
        excel_links = self._extract_excel_links(page_html)
        pdf_links = self._extract_visie_pdf_links(page_html)

        if len(excel_links) != 12:
            self.report["warnings"].append(
                f"Verwacht 12 Excel-doelensets, gevonden: {len(excel_links)}"
            )

        DATA_DIR.mkdir(parents=True, exist_ok=True)
        EXCEL_DIR.mkdir(parents=True, exist_ok=True)
        PDF_DIR.mkdir(parents=True, exist_ok=True)

        if not self.skip_download:
            self._download_excel_files(excel_links)
            self._download_pdf_files(pdf_links)
        else:
            logger.info("Download overgeslagen; bestaande bestanden worden hergebruikt.")

        records = self._parse_all_excel_files(excel_links)
        if not records:
            raise ValueError("Geen GO!-doelen uit Excel geëxtraheerd")

        self._write_output(records)

    def _fetch_page(self) -> str:
        logger.info("Pagina ophalen: %s", PAGE_URL)
        response = self.session.get(PAGE_URL, timeout=60)
        response.raise_for_status()
        return response.text

    def _extract_excel_links(self, html: str) -> list[dict[str, str]]:
        soup = BeautifulSoup(html, "html.parser")
        links: list[dict[str, str]] = []
        in_excel = False

        for element in soup.find_all(["button", "a"]):
            text = element.get_text(" ", strip=True)
            if text == "Doelensets Excel":
                in_excel = True
                continue
            if in_excel and text == "Doelensets PDF":
                break
            if not in_excel:
                continue
            if element.name != "a" or not element.get("href"):
                continue
            href = element["href"]
            if "/download/GOPRO-" not in href:
                continue
            label = clean_text(element.get_text(" ", strip=True))
            if not label.startswith("Doelenset BaO"):
                continue
            links.append(
                {
                    "label": label,
                    "url": urljoin(BASE_URL, href),
                    "discipline": discipline_from_label(label),
                }
            )

        # Dedupe op URL, behoud volgorde
        seen: set[str] = set()
        unique: list[dict[str, str]] = []
        for link in links:
            if link["url"] in seen:
                continue
            seen.add(link["url"])
            unique.append(link)

        for index, link in enumerate(unique, start=1):
            link["doelenset_nummer"] = str(index)

        logger.info("Excel-doelensets gevonden: %s", len(unique))
        return unique

    def _extract_visie_pdf_links(self, html: str) -> list[dict[str, str]]:
        soup = BeautifulSoup(html, "html.parser")
        links: list[dict[str, str]] = []
        seen: set[str] = set()

        for anchor in soup.find_all("a", href=True):
            label = clean_text(anchor.get_text(" ", strip=True))
            href = anchor["href"]
            if not label.startswith("Visietekst"):
                continue
            if "/download/GOPRO-" not in href:
                continue
            url = urljoin(BASE_URL, href)
            if url in seen:
                continue
            seen.add(url)
            links.append({"label": label, "url": url})

        logger.info("Visietekst-PDF's gevonden: %s", len(links))
        return links

    def _safe_filename(self, label: str, suffix: str) -> str:
        slug = re.sub(r"[^\w\s-]", "", label, flags=re.UNICODE)
        slug = re.sub(r"\s+", "_", slug).strip("_").casefold()
        return f"{slug}{suffix}"

    def _download_excel_files(self, links: list[dict[str, str]]) -> None:
        for link in links:
            filename = self._safe_filename(link["label"], ".xlsx")
            target = EXCEL_DIR / filename
            logger.info("Excel downloaden: %s", link["label"])
            response = self.session.get(link["url"], timeout=120)
            response.raise_for_status()
            target.write_bytes(response.content)
            link["local_path"] = str(target)
            self.report["excel_downloads"].append(
                {"label": link["label"], "path": str(target), "url": link["url"]}
            )

    def _download_pdf_files(self, links: list[dict[str, str]]) -> None:
        for link in links:
            filename = self._safe_filename(link["label"], ".pdf")
            target = PDF_DIR / filename
            logger.info("PDF downloaden: %s", link["label"])
            response = self.session.get(link["url"], timeout=120)
            response.raise_for_status()
            target.write_bytes(response.content)
            self.report["pdf_downloads"].append(
                {"label": link["label"], "path": str(target), "url": link["url"]}
            )

    def _resolve_excel_path(self, link: dict[str, str]) -> Path:
        if link.get("local_path"):
            return Path(link["local_path"])
        filename = self._safe_filename(link["label"], ".xlsx")
        return EXCEL_DIR / filename

    def _parse_all_excel_files(self, links: list[dict[str, str]]) -> list[dict[str, Any]]:
        records: list[dict[str, Any]] = []
        for link in links:
            path = self._resolve_excel_path(link)
            if not path.exists():
                raise FileNotFoundError(f"Excel-bestand ontbreekt: {path}")
            parsed = self._parse_excel_file(
                path,
                doelenset_nummer=int(link["doelenset_nummer"]),
                discipline=link["discipline"],
            )
            logger.info(
                "Geparst: %s → %s doelen",
                link["discipline"],
                len(parsed),
            )
            records.extend(parsed)
        return records

    def _parse_excel_file(
        self,
        path: Path,
        *,
        doelenset_nummer: int,
        discipline: str,
    ) -> list[dict[str, Any]]:
        workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
        records: list[dict[str, Any]] = []

        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
            header = [worksheet.cell(1, col).value for col in range(1, worksheet.max_column + 1)]
            leerjaar_columns: list[tuple[int, str]] = []
            for index, name in enumerate(header):
                if isinstance(name, str) and re.match(r"^\d", name.strip()):
                    leerjaar_columns.append((index, clean_text(name)))

            current_onderwerp = discipline_from_label(discipline) or clean_text(sheet_name)

            for row in worksheet.iter_rows(min_row=2, values_only=True):
                row_type = clean_text(row[0] if row else "")
                if row_type == "onderwerp" and len(row) > 2 and row[2]:
                    current_onderwerp = clean_text(row[2])
                    continue
                if row_type != "doelzin":
                    continue

                code = clean_text(row[1] if len(row) > 1 else "")
                titel = clean_text(row[2] if len(row) > 2 else "")
                if not code or not titel:
                    continue

                component = normalize_component(row[12] if len(row) > 12 else "")
                leerjaren = [
                    label
                    for col_index, label in leerjaar_columns
                    if col_index < len(row) and row[col_index] is True
                ]

                records.append(
                    {
                        "doelenset_nummer": doelenset_nummer,
                        "discipline": discipline_from_label(discipline),
                        "component": component,
                        "code": code,
                        "titel": titel,
                        "leerjaren": leerjaren,
                        "netwerk": "GO_NIEUW",
                    }
                )

        workbook.close()
        return records

    def _write_output(self, records: list[dict[str, Any]]) -> None:
        self.output.parent.mkdir(parents=True, exist_ok=True)
        with self.output.open("w", encoding="utf-8") as handle:
            for record in records:
                handle.write(json.dumps(record, ensure_ascii=False) + "\n")

        discipline_counts = dict(sorted(Counter(r["discipline"] for r in records).items()))
        self.report["discipline_counts"] = discipline_counts
        self.report["unique_goals"] = len(records)
        self.report["finished_at"] = datetime.now(timezone.utc).isoformat()
        self.report_path.write_text(
            json.dumps(self.report, ensure_ascii=False, indent=2) + "\n",
            encoding="utf-8",
        )

        metadata = {
            "network": "GO_NIEUW",
            "onderwijsniveau": "basisonderwijs",
            "brontitel": "GO! - Nieuw leerplan basisonderwijs (12 doelensets)",
            "source_url": PAGE_URL,
            "record_count": len(records),
            "format": "application/x-ndjson",
            "generated_at": datetime.now(timezone.utc).isoformat(),
        }
        self.metadata_path.write_text(
            json.dumps(metadata, ensure_ascii=False, indent=2) + "\n",
            encoding="utf-8",
        )

        logger.info("JSONL opgeslagen: %s", self.output)
        logger.info("Rapport opgeslagen: %s", self.report_path)
        logger.info("Eindresultaat: %s doelen", len(records))


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Download en parseer GO! nieuw leerplan doelensets."
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=OUTPUT_PATH,
        help=f"JSONL-uitvoerpad (standaard: {OUTPUT_PATH})",
    )
    parser.add_argument(
        "--skip-download",
        action="store_true",
        help="Sla downloads over en parseer bestaande Excel-bestanden.",
    )
    parser.add_argument("-v", "--verbose", action="store_true")
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    configure_logging(args.verbose)
    try:
        GoNieuwFetcher(
            output=args.output,
            skip_download=args.skip_download,
        ).run()
        return 0
    except KeyboardInterrupt:
        logger.error("Fetch onderbroken door gebruiker")
        return 130
    except Exception:
        logger.exception("GO! nieuw leerplan fetch mislukt")
        return 1


if __name__ == "__main__":
    sys.exit(main())
